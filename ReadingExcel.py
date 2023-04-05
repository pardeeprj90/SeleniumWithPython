import openpyxl
file_path = "E:\Data Analytics\Tableu Learning\Sample - Superstore.xlsx"
workbook = openpyxl.load_workbook(file_path)
orders_sheet = workbook["Orders"]
numberOfRows = orders_sheet.max_row
numberOfColumns = orders_sheet.max_column
for row in range (1,numberOfRows+1):
    for column in range(1,numberOfColumns+1):
        print(orders_sheet.cell(row, column).value,end='    ')
    print()

